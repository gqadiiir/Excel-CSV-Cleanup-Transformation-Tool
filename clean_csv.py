#!/usr/bin/env python3
"""
clean_csv.py — Professional Excel/CSV Cleanup & Transformation Tool
====================================================================
Accepts a messy CSV or Excel file and produces:
  1. A cleaned Excel workbook with formatted data
  2. A Summary sheet showing exactly what was fixed
  3. An Issues sheet listing every problem row and why it was flagged
  4. A console report showing the full cleanup breakdown

Problems detected and fixed automatically:
  - Duplicate rows (exact)
  - Inconsistent date formats -> normalised to YYYY-MM-DD
  - Inconsistent capitalisation (names, emails, status fields)
  - Whitespace padding in all cells
  - Invalid email addresses (flagged, not removed)
  - Missing required fields (flagged with clear reason)
  - Phone numbers normalised to (XXX) XXX-XXXX format
  - Salary/numeric columns stripped of non-numeric characters

Usage:
  python clean_csv.py input.csv
  python clean_csv.py input.xlsx --output cleaned.xlsx --required Email,Salary
  python clean_csv.py input.csv --no-color

Author  : [Your Name]
Version : 1.0.0
Requires: pandas, openpyxl  (pip install pandas openpyxl)
"""

import argparse
import re
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 1 — CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────

DEFAULT_REQUIRED_COLUMNS = ["Email"]
DEFAULT_DATE_COLUMNS     = ["Start Date", "Date", "DOB", "Hire Date"]
DEFAULT_NAME_COLUMNS     = ["First Name", "Last Name", "Full Name", "Name"]
DEFAULT_EMAIL_COLUMNS    = ["Email", "Email Address"]
DEFAULT_PHONE_COLUMNS    = ["Phone", "Phone Number", "Mobile", "Tel"]
DEFAULT_NUMERIC_COLUMNS  = ["Salary", "Amount", "Revenue", "Cost", "Price"]
DEFAULT_STATUS_COLUMNS   = ["Status", "Active", "State"]

COLOUR = {
    "header_bg" : "1D3557",
    "header_fg" : "FFFFFF",
    "summary_bg": "457B9D",
    "summary_fg": "FFFFFF",
    "ok_bg"     : "D4EDDA",
    "warn_bg"   : "FFF3CD",
    "crit_bg"   : "F8D7DA",
    "alt_row"   : "F1F8FF",
}

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 2 — CLEANING FUNCTIONS
# ─────────────────────────────────────────────────────────────────────────────

def strip_whitespace(df):
    issues = []
    for col in df.select_dtypes(include=["object", "str"]).columns:
        original = df[col].copy()
        df[col] = df[col].str.strip()
        changed = (original != df[col]) & original.notna()
        for idx in df[changed].index:
            issues.append({"Row": idx + 2, "Column": col, "Issue": "Whitespace stripped",
                           "Original": repr(original[idx]), "Fixed To": repr(df.at[idx, col])})
    return df, issues


def remove_duplicates(df):
    before = len(df)
    df = df.drop_duplicates().reset_index(drop=True)
    dropped = before - len(df)
    issues = []
    if dropped:
        issues.append({"Row": "Multiple", "Column": "All",
                       "Issue": f"{dropped} exact duplicate row(s) removed",
                       "Original": f"{dropped} rows", "Fixed To": "Removed"})
    return df, issues, dropped


def normalise_dates(df, date_cols):
    issues = []
    formats = ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y",
               "%Y/%m/%d", "%d %b %Y", "%B %d, %Y", "%d.%m.%Y"]
    for col in date_cols:
        if col not in df.columns:
            continue
        for idx, val in df[col].items():
            if pd.isna(val) or val == "":
                continue
            original = str(val).strip()
            parsed = None
            # Try explicit formats first to avoid ambiguity warnings
            for fmt in formats:
                try:
                    parsed = datetime.strptime(original, fmt)
                    break
                except ValueError:
                    continue
            # Fall back to pandas auto-parse if no explicit format matched
            if not parsed:
                try:
                    import warnings
                    with warnings.catch_warnings():
                        warnings.simplefilter("ignore")
                        parsed = pd.to_datetime(original, dayfirst=False)
                except Exception:
                    pass
            if parsed:
                normalised = pd.Timestamp(parsed).strftime("%Y-%m-%d")
                if normalised != original:
                    issues.append({"Row": idx + 2, "Column": col,
                                   "Issue": "Date format normalised",
                                   "Original": original, "Fixed To": normalised})
                    df.at[idx, col] = normalised
            else:
                issues.append({"Row": idx + 2, "Column": col,
                               "Issue": "Unrecognised date — review manually",
                               "Original": original, "Fixed To": "Not changed"})
    return df, issues


def normalise_names(df, name_cols):
    issues = []
    for col in name_cols:
        if col not in df.columns:
            continue
        for idx, val in df[col].items():
            if pd.isna(val) or val == "":
                continue
            original = str(val)
            fixed = original.title()
            if fixed != original:
                issues.append({"Row": idx + 2, "Column": col,
                               "Issue": "Name capitalisation fixed",
                               "Original": original, "Fixed To": fixed})
                df.at[idx, col] = fixed
    return df, issues


def normalise_emails(df, email_cols):
    pattern = re.compile(r"^[\w\.\+\-]+@[\w\-]+\.[a-zA-Z]{2,}$")
    issues = []
    for col in email_cols:
        if col not in df.columns:
            continue
        for idx, val in df[col].items():
            if pd.isna(val) or val == "":
                continue
            original = str(val).strip()
            lowered = original.lower()
            if lowered != original:
                issues.append({"Row": idx + 2, "Column": col,
                               "Issue": "Email lowercased",
                               "Original": original, "Fixed To": lowered})
                df.at[idx, col] = lowered
            if not pattern.match(lowered):
                issues.append({"Row": idx + 2, "Column": col,
                               "Issue": "Invalid email format — review manually",
                               "Original": lowered, "Fixed To": "Not changed"})
    return df, issues


def normalise_phones(df, phone_cols):
    issues = []
    for col in phone_cols:
        if col not in df.columns:
            continue
        for idx, val in df[col].items():
            if pd.isna(val) or val == "":
                continue
            original = str(val).strip()
            digits = re.sub(r"\D", "", original)
            if len(digits) == 10:
                formatted = f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
            elif len(digits) == 11 and digits[0] == "1":
                formatted = f"({digits[1:4]}) {digits[4:7]}-{digits[7:]}"
            else:
                formatted = original
            if formatted != original:
                issues.append({"Row": idx + 2, "Column": col,
                               "Issue": "Phone number formatted",
                               "Original": original, "Fixed To": formatted})
                df.at[idx, col] = formatted
    return df, issues


def normalise_numerics(df, numeric_cols):
    issues = []
    for col in numeric_cols:
        if col not in df.columns:
            continue
        for idx, val in df[col].items():
            if pd.isna(val) or val == "":
                continue
            original = str(val).strip()
            cleaned = re.sub(r"[^\d\.\-]", "", original)
            if not cleaned:
                continue
            try:
                numeric_val = float(cleaned)
                # Format: drop .0 for whole numbers, keep decimals otherwise
                formatted_val = str(int(numeric_val)) if numeric_val == int(numeric_val) \
                                else str(numeric_val)
                if formatted_val != original:
                    issues.append({"Row": idx + 2, "Column": col,
                                   "Issue": "Non-numeric characters stripped",
                                   "Original": original, "Fixed To": formatted_val})
                df.at[idx, col] = formatted_val
            except ValueError:
                issues.append({"Row": idx + 2, "Column": col,
                               "Issue": "Could not parse as number — review manually",
                               "Original": original, "Fixed To": "Not changed"})
    return df, issues


def normalise_status(df, status_cols):
    issues = []
    for col in status_cols:
        if col not in df.columns:
            continue
        for idx, val in df[col].items():
            if pd.isna(val) or val == "":
                continue
            original = str(val)
            fixed = original.strip().title()
            if fixed != original:
                issues.append({"Row": idx + 2, "Column": col,
                               "Issue": "Status capitalisation fixed",
                               "Original": original, "Fixed To": fixed})
                df.at[idx, col] = fixed
    return df, issues


def flag_missing_required(df, required_cols):
    issues = []
    for col in required_cols:
        if col not in df.columns:
            continue
        blank = df[col].isna() | (df[col].astype(str).str.strip() == "")
        for idx in df[blank].index:
            issues.append({"Row": idx + 2, "Column": col,
                           "Issue": f"Missing required field: {col}",
                           "Original": "", "Fixed To": "Flagged — manual review needed"})
    return issues


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 3 — EXCEL WRITER
# ─────────────────────────────────────────────────────────────────────────────

def _fill(hex_col):  return PatternFill("solid", fgColor=hex_col)
def _font(bold=False, colour="000000", size=11): return Font(bold=bold, color=colour, size=size)
def _border():
    s = Side(style="thin", color="DDDDDD")
    return Border(left=s, right=s, top=s, bottom=s)

def _header_row(ws, headers, bg, fg, row=1):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill      = _fill(bg)
        cell.font      = _font(bold=True, colour=fg)
        cell.border    = _border()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def _col_widths(ws, max_w=40):
    for col in ws.columns:
        w = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, max_w)


def write_excel_output(df_clean, issues, stats, output_path):
    wb = openpyxl.Workbook()

    # Sheet 1 — Cleaned Data
    ws = wb.active
    ws.title = "Cleaned Data"
    ws.freeze_panes = "A2"
    _header_row(ws, list(df_clean.columns), COLOUR["header_bg"], COLOUR["header_fg"])

    issue_rows = {i["Row"] for i in issues if isinstance(i["Row"], int)}

    for r, (_, row) in enumerate(df_clean.iterrows(), start=2):
        fill = _fill(COLOUR["warn_bg"]) if r in issue_rows else \
               (_fill(COLOUR["alt_row"]) if r % 2 == 0 else None)
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border    = _border()
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill

    _col_widths(ws)
    ws.row_dimensions[1].height = 28

    # Sheet 2 — Summary
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Cleanup Summary Report"
    ws2["A1"].font = _font(bold=True, colour=COLOUR["header_bg"], size=14)
    ws2["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws2["A2"].font = _font(colour="888888", size=10)

    rows = [
        ("Rows in original file",       stats["rows_original"]),
        ("Rows in cleaned file",         stats["rows_clean"]),
        ("Duplicate rows removed",       stats["duplicates_removed"]),
        ("Total issues detected",        stats["total_issues"]),
        ("Dates normalised",             stats["dates_fixed"]),
        ("Names corrected",              stats["names_fixed"]),
        ("Emails lowercased",            stats["emails_fixed"]),
        ("Phones formatted",             stats["phones_fixed"]),
        ("Numeric values cleaned",       stats["numerics_fixed"]),
        ("Status fields corrected",      stats["status_fixed"]),
        ("Missing required fields",      stats["missing_required"]),
        ("Invalid emails flagged",       stats["invalid_emails"]),
    ]
    _header_row(ws2, ["Metric", "Value"], COLOUR["summary_bg"], COLOUR["summary_fg"], row=4)
    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["B"].width = 20

    for r, (label, value) in enumerate(rows, start=5):
        ws2.cell(row=r, column=1, value=label).border = _border()
        vc = ws2.cell(row=r, column=2, value=value)
        vc.font      = _font(bold=True)
        vc.border    = _border()
        vc.alignment = Alignment(horizontal="center")
        vc.fill      = _fill(COLOUR["ok_bg"] if (isinstance(value, int) and value == 0)
                             else COLOUR["warn_bg"])

    # Sheet 3 — Issues Log
    ws3 = wb.create_sheet("Issues Log")
    ws3.freeze_panes = "A2"
    _header_row(ws3, ["Row", "Column", "Issue", "Original Value", "Fixed To"],
                COLOUR["header_bg"], COLOUR["header_fg"])

    for r, iss in enumerate(issues, start=2):
        vals = [iss.get("Row"), iss.get("Column"), iss.get("Issue"),
                iss.get("Original"), iss.get("Fixed To")]
        needs_review = "review" in str(iss.get("Issue", "")).lower()
        fill = _fill(COLOUR["crit_bg"] if needs_review else COLOUR["ok_bg"])
        for c, v in enumerate(vals, start=1):
            cell = ws3.cell(row=r, column=c, value=str(v))
            cell.fill      = fill
            cell.border    = _border()
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    _col_widths(ws3)
    ws3.row_dimensions[1].height = 28
    wb.save(output_path)


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 4 — CONSOLE REPORT
# ─────────────────────────────────────────────────────────────────────────────

def print_report(stats, output_path, use_colour=True):
    G = "\033[92m" if use_colour else ""
    Y = "\033[93m" if use_colour else ""
    R = "\033[91m" if use_colour else ""
    C = "\033[96m" if use_colour else ""
    B = "\033[1m"  if use_colour else ""
    E = "\033[0m"  if use_colour else ""

    print(f"\n{B}{C}{'═'*52}{E}")
    print(f"{B}{C}  CSV CLEANUP COMPLETE{E}")
    print(f"{B}{C}{'═'*52}{E}\n")

    def line(label, val, warn=False):
        col = (R if warn and val > 0 else G)
        mark = f"{col}✔{E}" if val == 0 else f"{Y}►{E}"
        print(f"  {mark}  {label:<37} {B}{col}{val}{E}")

    print(f"  {B}Input / Output{E}")
    print(f"  {'Original rows:':<39} {stats['rows_original']}")
    print(f"  {'Cleaned rows:':<39} {stats['rows_clean']}")
    dr_col = R if stats['duplicates_removed'] else G
    print(f"  {'Duplicates removed:':<39} {B}{dr_col}{stats['duplicates_removed']}{E}\n")

    print(f"  {B}Fixes Applied{E}")
    line("Dates normalised",        stats["dates_fixed"])
    line("Names corrected",         stats["names_fixed"])
    line("Emails lowercased",       stats["emails_fixed"])
    line("Phones formatted",        stats["phones_fixed"])
    line("Numeric values cleaned",  stats["numerics_fixed"])
    line("Status fields corrected", stats["status_fixed"])

    print(f"\n  {B}Needs Manual Review{E}")
    line("Missing required fields",  stats["missing_required"], warn=True)
    line("Invalid emails flagged",   stats["invalid_emails"],   warn=True)

    print(f"\n  {B}Total issues logged:{E}  {stats['total_issues']}")
    print(f"\n  {B}{G}Output:{E}  {output_path.resolve()}")
    print(f"{B}{C}{'═'*52}{E}\n")


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 5 — MAIN
# ─────────────────────────────────────────────────────────────────────────────

def load_file(path):
    if path.suffix.lower() == ".csv":
        return pd.read_csv(path, dtype=str, keep_default_na=False)
    elif path.suffix.lower() in (".xlsx", ".xls"):
        return pd.read_excel(path, dtype=str, keep_default_na=False)
    else:
        print(f"Unsupported file type: {path.suffix}")
        sys.exit(1)


def run(input_path, output_path, required_cols, use_colour):
    print(f"\nLoading: {input_path.name}...")
    df = load_file(input_path)
    rows_original = len(df)
    print(f"Loaded {rows_original} rows x {len(df.columns)} columns.")

    all_issues, dupes = [], 0

    date_cols    = [c for c in DEFAULT_DATE_COLUMNS    if c in df.columns]
    name_cols    = [c for c in DEFAULT_NAME_COLUMNS    if c in df.columns]
    email_cols   = [c for c in DEFAULT_EMAIL_COLUMNS   if c in df.columns]
    phone_cols   = [c for c in DEFAULT_PHONE_COLUMNS   if c in df.columns]
    numeric_cols = [c for c in DEFAULT_NUMERIC_COLUMNS if c in df.columns]
    status_cols  = [c for c in DEFAULT_STATUS_COLUMNS  if c in df.columns]

    df, iss = strip_whitespace(df);              all_issues += iss
    df, iss, dupes = remove_duplicates(df);      all_issues += iss
    df, iss = normalise_dates(df, date_cols);    all_issues += iss
    df, iss = normalise_names(df, name_cols);    all_issues += iss
    df, iss = normalise_emails(df, email_cols);  all_issues += iss
    df, iss = normalise_phones(df, phone_cols);  all_issues += iss
    df, iss = normalise_numerics(df, numeric_cols); all_issues += iss
    df, iss = normalise_status(df, status_cols); all_issues += iss
    all_issues += flag_missing_required(df, required_cols)

    def count(kw): return sum(1 for i in all_issues if kw.lower() in i["Issue"].lower())

    stats = {
        "rows_original":      rows_original,
        "rows_clean":         len(df),
        "duplicates_removed": dupes,
        "dates_fixed":        count("date format"),
        "names_fixed":        count("name capitalisation"),
        "emails_fixed":       count("email lowercase"),
        "phones_fixed":       count("phone number"),
        "numerics_fixed":     count("non-numeric"),
        "status_fixed":       count("status capitalisation"),
        "missing_required":   count("missing required"),
        "invalid_emails":     count("invalid email"),
        "total_issues":       len(all_issues),
    }

    print("Writing output workbook...")
    write_excel_output(df, all_issues, stats, output_path)
    print_report(stats, output_path, use_colour)


def main():
    parser = argparse.ArgumentParser(
        description="Clean and transform a messy CSV or Excel file.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python clean_csv.py employees.csv
  python clean_csv.py data.xlsx --output cleaned.xlsx
  python clean_csv.py report.csv --required Email,Phone,Salary --no-color
        """
    )
    parser.add_argument("input",     help="Input CSV or Excel file path.")
    parser.add_argument("--output",  "-o", help="Output .xlsx path. Default: <input>_cleaned.xlsx")
    parser.add_argument("--required","-r", default=",".join(DEFAULT_REQUIRED_COLUMNS),
                        help="Comma-separated required columns.")
    parser.add_argument("--no-color", action="store_true", help="Disable colour output.")
    args = parser.parse_args()

    input_path  = Path(args.input)
    if not input_path.exists():
        print(f"File not found: {input_path}")
        sys.exit(1)

    output_path   = Path(args.output) if args.output else \
                    input_path.parent / f"{input_path.stem}_cleaned.xlsx"
    required_cols = [c.strip() for c in args.required.split(",") if c.strip()]

    run(input_path, output_path, required_cols, not args.no_color)


if __name__ == "__main__":
    main()
